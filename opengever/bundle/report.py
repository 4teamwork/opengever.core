from Acquisition import aq_parent
from collections import OrderedDict
from datetime import datetime
from opengever.base.behaviors.base import IOpenGeverBase
from opengever.base.behaviors.translated_title import ITranslatedTitle
from opengever.base.behaviors.translated_title import ITranslatedTitleSupport
from opengever.base.interfaces import IReferenceNumber
from opengever.base.schemadump.config import DEFAULT_MANAGEABLE_ROLES
from opengever.base.schemadump.config import MANAGEABLE_ROLES_BY_TYPE
from opengever.base.schemadump.config import SHORTNAMES_BY_ROLE
from opengever.bundle.loader import BUNDLE_JSON_TYPES
from opengever.bundle.sections.constructor import BUNDLE_GUID_KEY
from opengever.bundle.sections.constructor import GEVER_SQL_TYPES_TO_MODEL
from opengever.dossier.behaviors.dossier import IDossier
from opengever.ogds.base.utils import get_current_admin_unit
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from os.path import basename
from plone import api
from sqlalchemy import inspect
from zope.annotation import IAnnotations
import logging


log = logging.getLogger(__name__)
log.setLevel(logging.INFO)


class DataCollector(object):
    """Collect report data from Plone.
    """

    def __init__(self, bundle):
        self.bundle = bundle

        au = get_current_admin_unit()
        self.au_abbreviation = au.abbreviation if au else ''

        self.catalog = api.portal.get_tool('portal_catalog')

    def __call__(self):
        data = {'metadata': {}, 'permissions': {}}

        portal_types = BUNDLE_JSON_TYPES.values()
        portal_types.append('ftw.mail.mail')

        for portal_type in portal_types:
            data['metadata'][portal_type] = []
            data['permissions'][portal_type] = []
            log.info("Collecting %s" % portal_type)

            if portal_type in GEVER_SQL_TYPES_TO_MODEL.keys():
                self.collect_sql_data(portal_type, data)
            else:
                self.collect_plone_data(portal_type, data)

        return data

    def collect_sql_data(self, portal_type, data):
        model, primary_key = GEVER_SQL_TYPES_TO_MODEL[portal_type]
        attributes = [col.key for col in inspect(model).mapper.column_attrs]
        query = model.query.filter(
            getattr(model, primary_key).in_(self.bundle.constructed_guids))
        for obj in query:
            item_info = {attr: getattr(obj, attr) for attr in attributes}
            data['metadata'][portal_type].append(item_info)

    def collect_plone_data(self, portal_type, data):
        brains = self.catalog.unrestrictedSearchResults(
            portal_type=portal_type,
            bundle_guid=tuple(self.bundle.constructed_guids))

        for brain in brains:
            obj = brain.getObject()
            guid = IAnnotations(obj).get(BUNDLE_GUID_KEY)

            item_info = self.get_item_metadata(obj, guid)
            data['metadata'][portal_type].append(item_info)

            if portal_type not in (
                    'opengever.document.document', 'ftw.mail.mail'):
                permission_info = self.get_permissions(obj, guid)
                if permission_info:
                    data['permissions'][portal_type].extend(
                        permission_info)

    def get_file_field(self, obj):
        if obj.portal_type == 'opengever.document.document':
            file_field = obj.file
        elif obj.portal_type == 'ftw.mail.mail':
            file_field = obj.message
        else:
            file_field = None

        return file_field

    def get_title(self, obj):
        if ITranslatedTitleSupport.providedBy(obj):
            return ITranslatedTitle(obj).title_de
        return obj.title

    def get_reference_number(self, obj):
        reference_adapter = IReferenceNumber(obj)
        refnum = reference_adapter.get_number()
        # Strip AdminUnit abbreviation
        refnum = refnum.replace(self.au_abbreviation, '').strip()
        return refnum

    def get_item_metadata(self, obj, guid):
        path = '/'.join(obj.getPhysicalPath())
        title = self.get_title(obj)
        parent = aq_parent(obj)
        parent_guid = IAnnotations(parent).get(BUNDLE_GUID_KEY)

        if obj.portal_type == 'opengever.repository.repositoryfolder':
            item_info = OrderedDict([
                ('guid', guid),
                ('parent_guid', parent_guid),
                ('path', path),
                ('title', title),
                ('description', obj.description),
                ('full_reference_number', self.get_reference_number(obj)),
            ])
        elif obj.portal_type == 'opengever.dossier.businesscasedossier':
            item_info = OrderedDict([
                ('guid', guid),
                ('parent_guid', parent_guid),
                ('path', path),
                ('title', title),
                ('full_reference_number', self.get_reference_number(obj)),
                ('reference_number', IDossier(obj).reference_number),
                ('responsible', IDossier(obj).responsible),
                ('description', IOpenGeverBase(obj).description),
                ('start', IDossier(obj).start),
                ('end', IDossier(obj).end),
                ('review_state', api.content.get_state(obj)),
            ])
        elif obj.portal_type in (
                'opengever.document.document', 'ftw.mail.mail'):
            file_field = self.get_file_field(obj)
            file_size = None
            file_name = None
            if file_field is not None:
                file_size = file_field.getSize()
                file_name = file_field.filename

            item_info = OrderedDict([
                ('guid', guid),
                ('parent_guid', parent_guid),
                ('path', path),
                ('title', title),
                ('file_size', file_size),
                ('file_name', file_name),
                ('document_date', obj.document_date),
            ])
        else:
            item_info = OrderedDict([
                ('guid', guid),
                ('path', path),
                ('title', title),
            ])

        return item_info

    def get_permissions(self, obj, guid):
        local_roles = obj.get_local_roles()
        inheritance_blocked = getattr(obj, '__ac_local_roles_block__', False)

        manageable_roles_for_type = MANAGEABLE_ROLES_BY_TYPE.get(
            obj.portal_type, DEFAULT_MANAGEABLE_ROLES)

        # Prepare a template to use as a base for rows, in order to make sure
        # we always have exactly the same columns in the same order
        row_template = OrderedDict([
            ('guid', guid),
            ('principal', None)]
            + [(role, None) for role in manageable_roles_for_type]
            + [('blocked_inheritance', None)],
        )

        # Always include at least one row per object with the info whether
        # or not inheritance is blocked
        inheritance_blocked_row = row_template.copy()
        inheritance_blocked_row['blocked_inheritance'] = inheritance_blocked

        # If local role assignments are present, include them as additional
        # rows (using GUID as key), one row per principal and their roles
        principal_role_rows = []
        for principal, roles in dict(local_roles).items():

            # Prepare row with principal name and False for all roles
            principal_roles = row_template.copy()
            principal_roles['principal'] = principal
            for role in manageable_roles_for_type:
                principal_roles[role] = False

            # Update with data from actual local roles
            for rolename in roles:
                if rolename == 'Owner':
                    continue

                short_role_name = SHORTNAMES_BY_ROLE[rolename]
                assert short_role_name in principal_roles
                principal_roles[short_role_name] = True

            principal_role_rows.append(principal_roles)

        return [inheritance_blocked_row] + principal_role_rows


class ASCIISummaryBuilder(object):
    """Build a quick ASCII summary of object counts based on report data.
    """

    def __init__(self, bundle):
        self.bundle = bundle
        self.report_data = bundle.report_data

    def build(self):
        report = []
        report.append('Imported objects:')
        report.append('=================')

        for portal_type, items in self.report_data['metadata'].items():
            line = "%s: %s" % (portal_type, len(items))
            report.append(line)

        return '\n'.join(report)


class XLSXReportBuilderBase(object):
    """Base class for XLSX report builders.
    """

    def __init__(self, *args, **kwargs):
        self.workbook = Workbook()
        self.workbook.remove(self.workbook.active)

    def build_and_save(self, report_path):
        self.write_report_data()
        self.save_report(report_path)

    def save_report(self, path):
        with open(path, 'w') as report_xlsx:
            self.workbook.save(report_xlsx)
            log.info("Wrote report to %s" % path)
        return path

    def add_sheet(self, sheet_name):
        log.info("Creating sheet %s" % sheet_name)
        sheet = self.workbook.create_sheet(sheet_name)
        sheet.title = sheet_name
        return sheet

    def write_row(self, sheet, rownum, values, bold=False, firstbold=False):
        for col_num, value in enumerate(values, 1):
            cell = sheet.cell(row=rownum + 1, column=col_num)
            cell.value = value
            if isinstance(value, datetime):
                sheet.column_dimensions[get_column_letter(col_num)].width = 20
            if bold or (firstbold and col_num == 1):
                cell.font = Font(bold=True)

    def write_report_data(self):
        raise NotImplementedError("To be implemented by subclasses")


class XLSXMainReportBuilder(XLSXReportBuilderBase):
    """Build a detailed report in XLSX format based on report data.
    """

    def __init__(self, bundle):
        super(XLSXMainReportBuilder, self).__init__()
        self.bundle = bundle

        report_data = self.bundle.report_data

        self.report_data = {'metadata': {}, 'permissions': {}}
        metadata = self.report_data['metadata']
        metadata.update(report_data['metadata'])
        permissions = self.report_data['permissions']
        permissions.update(report_data['permissions'])
        self.metadata = metadata
        self.permissions = permissions

        # Include mails in documents by moving them over
        docs = metadata.get('opengever.document.document', [])
        mails = metadata.get('ftw.mail.mail', [])
        metadata['opengever.document.document'] = docs + mails
        metadata.pop('ftw.mail.mail', None)

        docs = permissions.get('opengever.document.document', [])
        mails = permissions.get('ftw.mail.mail', [])
        permissions['opengever.document.document'] = docs + mails
        permissions.pop('ftw.mail.mail', None)

    def _write_summary(self):
        """Write a summary sheet with information regarding the entire import.
        """
        sheet_name = 'summary'
        sheet = self.add_sheet(sheet_name)

        stats = self.bundle.stats
        timings = stats['timings']
        duration = timings['migration_finished'] - timings['start_loading']
        bundle_name = basename(self.bundle.bundle_path.rstrip('/'))

        summary_stats = OrderedDict([
            ('bundle_name', bundle_name),
            ('start_time', timings['start_loading']),
            ('duration', duration),
        ])

        for rownum, item in enumerate(summary_stats.items()):
            stat_name, stat_value = item
            self.write_row(
                sheet, rownum, (stat_name, stat_value), firstbold=True)

    def _write_metadata(self):
        for json_name, portal_type in BUNDLE_JSON_TYPES.items():
            short_name = json_name.replace('.json', '')
            sheet = self.add_sheet(short_name)

            # Skip portal_types with no info
            if not self.metadata[portal_type]:
                continue

            # Label Row
            self.write_row(
                sheet, 0, self.metadata[portal_type][0].keys(), bold=True)

            # Data rows
            for rownum, info in enumerate(self.metadata[portal_type], 1):
                self.write_row(sheet, rownum, info.values())

    def _write_permissions(self):
        for json_name, portal_type in BUNDLE_JSON_TYPES.items():
            if not self.permissions.get(portal_type):
                continue

            sheet_name = '%s_permissions' % json_name.replace('.json', '')
            sheet = self.add_sheet(sheet_name)

            # Label Row
            headers = self.permissions[portal_type][0].keys()
            self.write_row(sheet, 0, headers, bold=True)

            # Data rows
            permission_infos = self.permissions[portal_type]
            for rownum, perm_info in enumerate(permission_infos, 1):
                self.write_row(sheet, rownum, perm_info.values())

    def write_report_data(self):
        self._write_summary()
        self._write_metadata()
        self._write_permissions()


class XLSXValidationReportBuilder(XLSXReportBuilderBase):
    """Build a validation report in XLSX format based on `errors` and
    `warnings` dictionaries.
    """

    ERROR_FIELDS = OrderedDict([
        ('files_not_found', ('guid', 'filepath', 'ogg_path')),
        ('files_io_errors', ('guid', 'filepath', 'ioerror', 'ogg_path')),
        ('files_unresolvable_path', ('guid', 'filepath', 'ogg_path')),
        ('files_invalid_types', ('guid', 'filepath', 'ogg_path')),
        ('unmapped_unc_mounts', ('mount', )),
    ])

    WARNING_FIELDS = OrderedDict([
        ('max_nesting_depth_exceeded', ('guid', 'max', 'actual', 'path')),
    ])

    def __init__(self, bundle):
        super(XLSXValidationReportBuilder, self).__init__()
        self.bundle = bundle
        self.errors = bundle.errors
        self.warnings = bundle.warnings

    def write_summary(self):
        """Write a summary with counts for every message type.
        """
        sheet_name = 'summary'
        sheet = self.add_sheet(sheet_name)

        rownum = 0
        self.write_row(sheet, rownum, ['Errors'], bold=True)
        rownum += 1

        for error_type, error_list in self.errors.items():
            self.write_row(
                sheet, rownum, (error_type, len(error_list)),
                firstbold=True)
            rownum += 1

        rownum += 1
        self.write_row(sheet, rownum, ['Warnings'], bold=True)
        rownum += 1

        for warning_type, warning_list in self.warnings.items():
            self.write_row(
                sheet, rownum, (warning_type, len(warning_list)),
                firstbold=True)
            rownum += 1

    def write_msg_sheets(self, msg_dict, field_defs):
        """Write a sheet for every message type in msg_dict.
        """
        for msg_type, msg_list in msg_dict.items():
            try:
                fields = field_defs[msg_type]
            except KeyError:
                log.warn('Unknown message type %r, skipping.' % msg_type)
                continue

            sheet = self.add_sheet(msg_type)

            # Label Row
            self.write_row(sheet, 0, fields, bold=True)

            # Data rows
            for rownum, msg in enumerate(msg_list, 1):
                self.write_row(sheet, rownum, msg)

    def write_report_data(self):
        self.write_summary()
        self.write_msg_sheets(self.errors, self.ERROR_FIELDS)
        self.write_msg_sheets(self.warnings, self.WARNING_FIELDS)
