from datetime import datetime
from ftw.testbrowser import browsing
from ftw.testing import freeze
from opengever.testing import SolrIntegrationTestCase
from openpyxl import load_workbook
from tempfile import NamedTemporaryFile
import os
import pytz


class TestExcelRoleAssignmentReport(SolrIntegrationTestCase):

    @browsing
    def test_role_assignment_report(self, browser):
        self.login(self.administrator, browser=browser)

        url = u'{}/download-role-assignment-report'.format(
            self.portal.absolute_url())

        with freeze(datetime(2017, 10, 16, 0, 0, tzinfo=pytz.utc)):
            browser.open(url, view="?filters.principal_ids:record:list=jurgen.konig")

        self.assertEquals(
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            browser.headers['content-type'])

        data = browser.contents
        with NamedTemporaryFile(delete=False, suffix='.xlsx') as tmpfile:
            tmpfile.write(data)
            tmpfile.flush()
            workbook = load_workbook(tmpfile.name)
            os.remove(tmpfile.name)

        rows = list(workbook.active.rows)

        self.assertSequenceEqual(
            [
                ["Title", "URL", "Portal type", "Principal id", "User name", "Group name", "Role"],
                [
                    "Ordnungssystem",
                    "http://nohost/plone/ordnungssystem",
                    "Repository Root",
                    "jurgen.konig",
                    "jurgen.konig",
                    None,
                    "Reviewer",
                ],
                [
                    "Ordnungssystem",
                    "http://nohost/plone/ordnungssystem",
                    "Repository Root",
                    "jurgen.konig",
                    "jurgen.konig",
                    None,
                    "Reactivate",
                ],
            ],
            [[cell.value for cell in row] for row in rows])

    @browsing
    def test_role_assignment_report_group_filter(self, browser):
        self.login(self.administrator, browser=browser)

        url = u'{}/download-role-assignment-report'.format(
            self.portal.absolute_url())

        with freeze(datetime(2017, 10, 16, 0, 0, tzinfo=pytz.utc)):
            browser.open(url, view="?filters.principal_ids:record:list=group:fa_inbox_users")

        self.assertEquals(
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            browser.headers['content-type'])

        data = browser.contents
        with NamedTemporaryFile(delete=False, suffix='.xlsx') as tmpfile:
            tmpfile.write(data)
            tmpfile.flush()
            workbook = load_workbook(tmpfile.name)
            os.remove(tmpfile.name)

        rows = list(workbook.active.rows)

        self.assertSequenceEqual(
            [
                ["Title", "URL", "Portal type", "Principal id", "User name", "Group name", "Role"],
                [
                    u'Vertr\xe4ge mit der kantonalen Finanzverwaltung',
                    u'http://nohost/plone/ordnungssystem/fuhrung/vertrage-und-vereinbarungen/dossier-1',
                    u'Business Case Dossier',
                    u'fa_inbox_users',
                    None,
                    u'fa_inbox_users',
                    u'Task responsible'
                ],
                [
                    u'Zu allem \xdcbel',
                    u'http://nohost/plone/ordnungssystem/fuhrung/vertrage-und-vereinbarungen/dossier-17',
                    u'Business Case Dossier',
                    u'fa_inbox_users',
                    None,
                    u'fa_inbox_users',
                    u'Task responsible'
                ],
                [
                    u'Abgeschlossene Vertr\xe4ge',
                    u'http://nohost/plone/ordnungssystem/fuhrung/vertrage-und-vereinbarungen/dossier-5',
                    u'Business Case Dossier',
                    u'fa_inbox_users',
                    None,
                    u'fa_inbox_users',
                    u'Task responsible'
                ],
                [
                    u'Inaktive Vertr\xe4ge',
                    u'http://nohost/plone/ordnungssystem/fuhrung/vertrage-und-vereinbarungen/dossier-6',
                    u'Business Case Dossier',
                    u'fa_inbox_users',
                    None,
                    u'fa_inbox_users',
                    u'Task responsible'
                ]
            ],
            [[cell.value for cell in row] for row in rows])

    @browsing
    def test_filename_includes_date_time(self, browser):
        self.login(self.administrator, browser=browser)

        url = u'{}/download-role-assignment-report'.format(
            self.portal.absolute_url())

        with freeze(datetime(2017, 10, 16, 0, 0, tzinfo=pytz.utc)):
            browser.open(url)

            self.assertEquals(
                'attachment; filename="role_assignment_report_2017-10-16_00-00.xlsx"',
                browser.headers['content-disposition'])

    @browsing
    def test_filename_includes_filter_information(self, browser):
        self.login(self.administrator, browser=browser)

        url = u'{}/download-role-assignment-report'.format(
            self.portal.absolute_url())

        with freeze(datetime(2017, 10, 16, 0, 0, tzinfo=pytz.utc)):
            browser.open(
                url,
                view="?filters.principal_ids:record:list=jurgen.konig"
                     "&filters.include_memberships:record:boolean=true&"
                     "&filters.root:record={}".format(
                         self.dossier.UID()))

        self.assertEquals(
            'attachment; filename="role_assignment_report_2017-10-16_00-00'
            '_branch_dossier-1'
            '_including_memberships_jurgen.konig.xlsx"',
            browser.headers['content-disposition'])
