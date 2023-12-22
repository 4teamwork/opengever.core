from datetime import datetime
from ftw.testbrowser import browsing
from io import BytesIO
from opengever.testing import SolrIntegrationTestCase
from openpyxl import load_workbook


class TestProposalReporter(SolrIntegrationTestCase):

    def load_workbook(self, data):
        return load_workbook(BytesIO(data))

    @browsing
    def test_empty_proposal_report(self, browser):
        self.login(self.regular_user, browser=browser)
        browser.open(view='proposal_report', data={'paths:list': []})

        self.assertEquals('Error You have not selected any items.',
                          browser.css('.portalMessage.error').text[0])

    @browsing
    def test_proposal_report(self, browser):
        self.login(self.regular_user, browser=browser)

        data = self.make_path_param(self.proposal, self.draft_proposal)
        data.update({'sort_on': 'sequence_number', 'sort_order': 'asc'})
        browser.open(view='proposal_report', data=data)

        workbook = self.load_workbook(browser.contents)

        # One title row and two data rows
        self.assertEquals(3, len(list(workbook.active.rows)))

        self.assertSequenceEqual(
            [u'Vertr\xe4ge',
             datetime(2016, 8, 31, 14, 9, 33),
             u'Submitted',
             u'Ziegler Robert (robert.ziegler)',
             u'Rechnungspr\xfcfungskommission',
             u'Vertr\xe4ge mit der kantonalen Finanzverwaltung',
             u'F\xfcr weitere Bearbeitung bewilligen'],
            [cell.value for cell in list(workbook.active.rows)[1]])

        self.assertSequenceEqual(
            [u'Antrag f\xfcr Kreiselbau',
             datetime(2016, 8, 31, 14, 13, 33),
             u'Active',
             u'Ziegler Robert (robert.ziegler)',
             u'Kommission f\xfcr Verkehr',
             u'Vertr\xe4ge mit der kantonalen Finanzverwaltung',
             None],
            [cell.value for cell in list(workbook.active.rows)[2]])
