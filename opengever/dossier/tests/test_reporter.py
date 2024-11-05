from datetime import date
from datetime import datetime
from ftw.builder import Builder
from ftw.builder import create
from ftw.testbrowser import browsing
from io import BytesIO
from opengever.dossier.behaviors.customproperties import IDossierCustomProperties
from opengever.dossier.behaviors.dossier import IDossier
from opengever.dossier.behaviors.filing import IFilingNumber
from opengever.dossier.filing.report import filing_no_filing
from opengever.dossier.filing.report import filing_no_number
from opengever.dossier.filing.report import filing_no_year
from opengever.testing import SolrIntegrationTestCase
from openpyxl import load_workbook
from zope.component import getMultiAdapter
import json


class TestDossierReporter(SolrIntegrationTestCase):

    def load_workbook(self, data):
        return load_workbook(BytesIO(data))

    def create_propertysheet_for(self, dossier):
        IDossier(dossier).dossier_type = u"businesscase"
        choices = [u'Rot', u'Gr\xfcn', u'Blau']
        create(
            Builder("property_sheet_schema")
            .named("schema1")
            .assigned_to_slots(u"IDossier.dossier_type.businesscase")
            .with_field("bool", u"yesorno", u"Yes or no", u"", False)
            .with_field("choice", u"color", u"Favorite Color", u"", False, values=choices)
            .with_field("multiple_choice", u"colors", u"Colors", u"", False, values=choices)
            .with_field("int", u"age", u"Age", u"", False)
            .with_field("text", u"poem", u"Poem", u"", False)
            .with_field("textline", u"tagline", u"Tag Line", u"", False)
            .with_field("date", u"birthday", u"Birthday", u"", False)
        )

    @browsing
    def test_dossier_report(self, browser):
        self.login(self.regular_user, browser=browser)

        browser.open(view='dossier_report',
                     # /plone/ordnungssystem/...
                     data=self.make_path_param(
                        self.dossier, self.inactive_dossier))

        workbook = self.load_workbook(browser.contents)
        rows = list(workbook.active.rows)

        self.assertSequenceEqual(
            [self.dossier.title,
             datetime(2016, 1, 1),
             None,
             u'Ziegler Robert (robert.ziegler)',
             u'Active',
             u'Client1 1.1 / 1'],
            [cell.value for cell in rows[1]])

        self.assertSequenceEqual(
            [self.inactive_dossier.title,
             datetime(2016, 1, 1, 0, 0),
             datetime(2016, 12, 31, 0, 0),
             u'Ziegler Robert (robert.ziegler)',
             u'Inactive',
             u'Client1 1.1 / 3'],
            [cell.value for cell in rows[2]])

    @browsing
    def test_sorts_rows_according_to_the_sort_params_in_request(self, browser):
        self.login(self.regular_user, browser=browser)

        paths1 = self.make_path_param(self.dossier, self.inactive_dossier)
        paths1.update({'sort_on': 'reference', 'sort_order': 'asc'})
        browser.open(view='dossier_report', data=paths1)

        workbook = self.load_workbook(browser.contents)
        rows = list(workbook.active.rows)

        self.assertEqual(self.dossier.title, rows[1][0].value)
        self.assertEqual(self.inactive_dossier.title, rows[2][0].value)

        paths2 = self.make_path_param(self.inactive_dossier, self.dossier)
        paths2.update({'sort_on': 'reference', 'sort_order': 'descending'})
        browser.open(view='dossier_report', data=paths2)

        workbook = self.load_workbook(browser.contents)
        rows = list(workbook.active.rows)

        self.assertEqual(self.inactive_dossier.title, rows[1][0].value)
        self.assertEqual(self.dossier.title, rows[2][0].value)

    @browsing
    def test_supports_explicit_columns_from_request(self, browser):
        self.login(self.regular_user, browser=browser)

        params = self.make_path_param(self.dossier)
        params.update({'columns': [
            'title',
            'start',
            'end',
            'external_reference',
            'responsible',
            'review_state',
            'reference',
            'touched',
            'keywords',
            'description',
            'creator'
        ]})
        browser.open(view='dossier_report', data=params)

        workbook = self.load_workbook(browser.contents)
        rows = list(workbook.active.rows)

        expected_titles = [
            u'Title',
            u'Start date',
            u'End date',
            u'External reference',
            u'Responsible',
            u'Review state',
            u'Reference number',
            u'Last modified',
            u'Keywords',
            u'Description',
            u'Creator'
        ]
        self.assertEqual(expected_titles, [cell.value for cell in rows[0]])

        expected_values = [
            self.dossier.title,
            datetime(2016, 1, 1),
            None,
            u'qpr-900-9001-\xf7',
            u'Ziegler Robert (robert.ziegler)',
            u'Active',
            u'Client1 1.1 / 1',
            datetime(2016, 8, 31, 0, 0),
            u'Finanzverwaltung, Vertr\xe4ge',
            self.dossier.description,
            u'Ziegler Robert (robert.ziegler)'
        ]
        self.assertEqual(expected_values, [cell.value for cell in rows[1]])

    @browsing
    def test_necessary_columns_are_aliased(self, browser):
        self.login(self.regular_user, browser=browser)

        params = self.make_path_param(self.dossier)

        # The frontend uses these names for the 'responsible' and
        # 'review_state' columns. They should be aliased accordingly in the
        # DossierReporter column_settings
        params.update({'columns': [
            'responsible_fullname',
            'review_state_label',
        ]})
        browser.open(view='dossier_report', data=params)

        workbook = self.load_workbook(browser.contents)
        rows = list(workbook.active.rows)

        expected_titles = [
            u'Responsible',
            u'Review state',
        ]
        self.assertEqual(expected_titles, [cell.value for cell in rows[0]])

        expected_values = [
            u'Ziegler Robert (robert.ziegler)',
            u'Active',
        ]
        self.assertEqual(expected_values, [cell.value for cell in rows[1]])

    @browsing
    def test_supports_custom_fields(self, browser):
        self.login(self.regular_user, browser=browser)

        self.create_propertysheet_for(self.dossier)
        IDossierCustomProperties(self.dossier).custom_properties = {
            "IDossier.dossier_type.businesscase": {
                "yesorno": True,
                "color": u'Gr\xfcn',
                "colors": [u'Rot', u'Gr\xfcn'],
                "age": 42,
                # Multiline Text fields are not currently indexed in Solr
                "poem": "Lorem\nIpsum",
                "tagline": "Woosh!",
                "birthday": date(2022, 4, 1),
            }
        }
        self.dossier.reindexObject()
        self.commit_solr()

        params = self.make_path_param(self.dossier)
        params.update({'columns': [
            'yesorno_custom_field_boolean',
            'color_custom_field_string',
            'colors_custom_field_strings',
            'age_custom_field_int',
            'tagline_custom_field_string',
            'birthday_custom_field_date',
        ]})
        browser.open(view='dossier_report', data=params)

        workbook = self.load_workbook(browser.contents)
        rows = list(workbook.active.rows)

        expected_titles = [
            u'Yes or no',
            u'Favorite Color',
            u'Colors',
            u'Age',
            u'Tag Line',
            u'Birthday',
        ]
        self.assertEqual(expected_titles, [cell.value for cell in rows[0]])

        expected_values = [
            True,
            u'Gr\xfcn',
            u'Gr\xfcn, Rot',
            42,
            u'Woosh!',
            datetime(2022, 4, 1, 0, 0),
        ]
        self.assertEqual(expected_values, [cell.value for cell in rows[1]])

    @browsing
    def test_sets_number_format_for_date_fields(self, browser):
        self.login(self.regular_user, browser=browser)

        self.create_propertysheet_for(self.dossier)
        IDossierCustomProperties(self.dossier).custom_properties = {
            "IDossier.dossier_type.businesscase": {
                "birthday": date(2022, 4, 1),
            }
        }
        IDossier(self.dossier).end = date(2033, 1, 1)
        self.dossier.reindexObject()
        self.commit_solr()

        params = self.make_path_param(self.dossier)
        params.update({'columns': [
            'start',
            'end',
            'touched',
            'birthday_custom_field_date',
        ]})
        browser.open(view='dossier_report', data=params)

        workbook = self.load_workbook(browser.contents)
        rows = list(workbook.active.rows)

        expected_values = [
            datetime(2016, 1, 1, 0, 0),
            datetime(2033, 1, 1, 0, 0),
            datetime(2016, 8, 31, 0, 0),
            datetime(2022, 4, 1, 0, 0),
        ]
        self.assertEqual(expected_values, [cell.value for cell in rows[1]])
        self.assertEqual(
            ['DD.MM.YYYY', 'DD.MM.YYYY', 'DD.MM.YYYY', 'DD.MM.YYYY'],
            [cell.number_format for cell in rows[1]],
        )

    @browsing
    def test_dossier_report_with_pseudorelative_path(self, browser):
        self.login(self.regular_user, browser=browser)

        browser.open(view='dossier_report',
                     # /ordnungssystem/...
                     data=self.make_pseudorelative_path_param(
                        self.dossier, self.inactive_dossier))

        workbook = self.load_workbook(browser.contents)
        rows = list(workbook.active.rows)

        self.assertSequenceEqual(
            [self.dossier.title,
             datetime(2016, 1, 1),
             None,
             u'Ziegler Robert (robert.ziegler)',
             u'Active',
             u'Client1 1.1 / 1'],
            [cell.value for cell in rows[1]])

        self.assertSequenceEqual(
            [self.inactive_dossier.title,
             datetime(2016, 1, 1, 0, 0),
             datetime(2016, 12, 31, 0, 0),
             u'Ziegler Robert (robert.ziegler)',
             u'Inactive',
             u'Client1 1.1 / 3'],
            [cell.value for cell in rows[2]])

    @browsing
    def test_respects_column_tabbedview_settings_if_exists(self, browser):
        self.login(self.regular_user, browser=browser)

        data = {'view_name': 'mydossiers',
                'gridstate': json.dumps({
                    "columns": [
                        {"id": "path_checkbox", "width": 30},
                        {"id": "sortable_title", "width": 300, "sortable": True},
                        {"id": "review_state", "width": 110, "sortable": True},
                        {"id": "reference", "width": 110, "sortable": True},
                        {"id": "end", "width": 110, "sortable": True},
                        {"id": "responsible", "width": 110, "sortable": True},
                        {"id": "start", "width": 110, "hidden": True, "sortable": True},
                        {"id": "Subject", "width": 110, "sortable": True},
                        {"id": "dummy", "width": 1, "hidden": True}],
                    "sort": {"field": "modified", "direction": "ASC"}})}
        browser.open(view='@@tabbed_view/setgridstate', data=data)

        data = self.make_path_param(self.dossier, self.inactive_dossier)
        data['view_name'] = 'mydossiers'
        browser.open(view='dossier_report', data=data)

        workbook = self.load_workbook(browser.contents)
        rows = list(workbook.active.rows)
        self.assertSequenceEqual(
            [u'Title', u'Review state', u'Reference number', u'End date',
             u'Responsible'],
            [cell.value for cell in rows[0]])

    @browsing
    def test_report_appends_filing_fields(self, browser):
        self.activate_feature('filing_number')
        self.login(self.regular_user, browser=browser)

        IFilingNumber(self.dossier).filing_no = u'Client1-Leitung-2012-1'
        self.dossier.reindexObject()
        self.commit_solr()

        browser.open(view='dossier_report',
                     data=self.make_path_param(self.dossier))

        workbook = self.load_workbook(browser.contents)

        labels = [cell.value for cell in list(workbook.active.rows)[0]]
        self.assertIn(u'Filing', labels)
        self.assertIn(u'Filing year', labels)
        self.assertIn(u'Filing number', labels)
        self.assertIn(u'Filing number', labels)

        self.assertSequenceEqual(
            [self.dossier.title,
             datetime(2016, 1, 1),
             None,
             u'Ziegler Robert (robert.ziegler)',
             u'Leitung',
             2012,
             1,
             u'Client1-Leitung-2012-1',
             u'Active',
             u'Client1 1.1 / 1'],
            [cell.value for cell in list(workbook.active.rows)[1]])

    def test_filing_no_year(self):
        self.assertEquals(
            filing_no_year('Client1-Leitung-2012-1'), 2012)
        self.assertEquals(
            filing_no_year('Leitung'), None)
        self.assertEquals(
            filing_no_year('Client1-Direktion-2011-555'), 2011)
        self.assertEquals(
            filing_no_year(None), None)

    def test_filing_no_number(self):
        self.assertEquals(
            filing_no_number('Client1-Leitung-2012-1'), 1)
        self.assertEquals(
            filing_no_number('Leitung'), None)
        self.assertEquals(
            filing_no_number('Client1-Direktion-2011-555'), 555)
        self.assertEquals(
            filing_no_number(None), None)

    def test_filing_no_filing(self):
        self.assertEquals(
            filing_no_filing('Client1-Leitung-2012-1'), 'Leitung')
        self.assertEquals(
            filing_no_filing('Leitung'), 'Leitung')
        self.assertEquals(
            filing_no_filing('Client1-Direktion-2011-555'), 'Direktion')
        self.assertEquals(
            filing_no_filing(None), None)

    @browsing
    def test_export_multiple_batches(self, browser):
        self.login(self.regular_user, browser=browser)

        params = self.make_path_param(self.dossier)

        # The frontend uses these names for the 'responsible' and
        # 'review_state' columns. They should be aliased accordingly in the
        # DossierReporter column_settings
        params.update({'columns': [
            'responsible_fullname',
            'review_state_label',
        ]})

        objs = [self.dossier, self.inactive_dossier]
        self.request['paths'] = ['/'.join(obj.getPhysicalPath()) for obj in objs]
        view = getMultiAdapter((self.portal, self.request), name='dossier_report')
        view.batch_size = 1

        workbook = self.load_workbook(view())
        rows = list(workbook.active.rows)

        self.assertEqual(1 + len(objs), len(rows)) # One additional row for the header

    @browsing
    def test_supports_query_by_listing(self, browser):
        self.login(self.regular_user, browser=browser)

        payload = {'listing_name': 'dossiers'}

        # Report
        browser.open(view='dossier_report', data=payload)
        workbook = self.load_workbook(browser.contents)
        rows = list(workbook.active.rows)

        # Listing
        view = '@listing?name=dossiers'
        browser.open(self.portal, view=view, headers=self.api_headers)

        report_titles = [row[0].value for row in rows[1:]]
        listing_titles = [item.get('title') for item in browser.json.get('items')]

        self.assertEqual(19, len(report_titles))
        self.assertEqual(listing_titles, report_titles)

    @browsing
    def test_supports_query_by_listing_with_filtering(self, browser):
        self.login(self.regular_user, browser=browser)

        payload = {
            'listing_name': 'dossiers',
            'filters.responsible:record:list': self.dossier_responsible.id
        }

        # Report
        browser.open(view='dossier_report', data=payload)
        workbook = self.load_workbook(browser.contents)
        rows = list(workbook.active.rows)

        # Listing
        view = '@listing?name=dossiers&filters.responsible:record:list={}'.format(
            self.dossier_responsible.id)
        browser.open(self.portal, view=view, headers=self.api_headers)

        report_titles = [row[0].value for row in rows[1:]]
        listing_titles = [item.get('title') for item in browser.json.get('items')]

        self.assertEqual(10, len(report_titles))
        self.assertEqual(listing_titles, report_titles)

    @browsing
    def test_supports_query_by_listing_with_searching(self, browser):
        self.login(self.regular_user, browser=browser)

        payload = {
            'listing_name': 'dossiers',
            'search': '"{}"'.format(self.dossier.title.encode('utf8'))
        }

        # Report
        browser.open(view='dossier_report', data=payload)
        workbook = self.load_workbook(browser.contents)
        rows = list(workbook.active.rows)

        # Listing
        view = '@listing?name=dossiers&search="{}"'.format(
            self.dossier.title.encode('utf-8'))
        browser.open(self.portal, view=view, headers=self.api_headers)

        report_titles = [row[0].value for row in rows[1:]]
        listing_titles = [item.get('title') for item in browser.json.get('items')]

        self.assertEqual(1, len(report_titles))
        self.assertEqual(listing_titles, report_titles)

    @browsing
    def test_supports_query_by_listing_with_ordering(self, browser):
        self.login(self.regular_user, browser=browser)

        payload = {
            'listing_name': 'dossiers',
            'sort_order': 'descending',
            'sort_on': 'is_subdossier'
        }

        # Report
        browser.open(view='dossier_report', data=payload)
        workbook = self.load_workbook(browser.contents)
        rows = list(workbook.active.rows)

        # Listing sorting descending
        view = '@listing?name=dossiers&sort_order=descending&sort_on=is_subdossier'
        browser.open(self.portal, view=view, headers=self.api_headers)
        listing_titles_descending = [item.get('title') for item in browser.json.get('items')]

        # Listing sorting ascending
        view = '@listing?name=dossiers&sort_order=ascending&sort_on=is_subdossier'
        browser.open(self.portal, view=view, headers=self.api_headers)
        listing_titles_ascending = [item.get('title') for item in browser.json.get('items')]

        report_titles = [row[0].value for row in rows[1:]]

        self.assertEqual(listing_titles_descending, report_titles)
        self.assertNotEqual(listing_titles_ascending, report_titles)
