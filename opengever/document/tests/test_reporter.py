from datetime import date
from datetime import datetime
from ftw.builder import Builder
from ftw.builder import create
from ftw.testbrowser import browsing
from io import BytesIO
from opengever.document.behaviors.customproperties import IDocumentCustomProperties
from opengever.document.behaviors.metadata import IDocumentMetadata
from opengever.testing import SolrIntegrationTestCase
from openpyxl import load_workbook
import json


class TestDocumentReporter(SolrIntegrationTestCase):

    def load_workbook(self, data):
        return load_workbook(BytesIO(data))

    def create_propertysheet_for(self, document):
        IDocumentMetadata(document).document_type = u"offer"
        choices = [u'Rot', u'Gr\xfcn', u'Blau']
        create(
            Builder("property_sheet_schema")
            .named("schema1")
            .assigned_to_slots(u"IDocumentMetadata.document_type.offer")
            .with_field("bool", u"yesorno", u"Yes or no", u"", False)
            .with_field("choice", u"color", u"Favorite Color", u"", False, values=choices)
            .with_field("multiple_choice", u"colors", u"Colors", u"", False, values=choices)
            .with_field("int", u"age", u"Age", u"", False)
            .with_field("text", u"poem", u"Poem", u"", False)
            .with_field("textline", u"tagline", u"Tag Line", u"", False)
            .with_field("date", u"birthday", u"Birthday", u"", False)
        )

    @browsing
    def test_empty_document_report(self, browser):
        self.login(self.regular_user, browser=browser)
        browser.open(view='document_report', data={'paths:list': []})

        self.assertEquals('Error You have not selected any Items',
                          browser.css('.portalMessage.error').text[0])

    @browsing
    def test_document_report(self, browser):
        self.login(self.regular_user, browser=browser)

        # /plone/ordnungssystem/...
        data = self.make_path_param(self.document, self.mail_eml)
        data['sort_on'] = 'sequence_number'
        data['sort_order'] = 'desc'

        browser.open(view='document_report', data=data)

        workbook = self.load_workbook(browser.contents)

        # One title row and two data rows
        self.assertEquals(3, len(list(workbook.active.rows)))

        self.assertSequenceEqual(
            [u'Client1 1.1 / 1 / 14',
             14,
             u'Vertr\xe4gsentwurf',
             u'test_user_1_',
             datetime(2010, 1, 3, 0, 0),
             datetime(2010, 1, 3, 0, 0),
             datetime(2010, 1, 3, 0, 0),
             None,
             u'not assessed',
             u'Vertr\xe4ge mit der kantonalen Finanzverwaltung'],
            [cell.value for cell in list(workbook.active.rows)[1]])

        self.assertSequenceEqual(
            [u'Client1 1.1 / 1 / 29',
             29,
             u'Die B\xfcrgschaft',
             u'Freddy H\xf6lderlin <from@example.org>',
             datetime(1999, 1, 1, 0, 0),
             datetime(2016, 8, 31, 0, 0),
             None,
             None,
             u'not assessed',
             u'Vertr\xe4ge mit der kantonalen Finanzverwaltung'],
            [cell.value for cell in list(workbook.active.rows)[2]])

    @browsing
    def test_supports_explicit_columns_from_request(self, browser):
        self.login(self.regular_user, browser=browser)

        params = self.make_path_param(self.document)
        params.update({'columns': [
            'title',
            'reference',
            'changed',
            'document_author',
            'containing_dossier',
            'containing_subdossier',
            'sequence_number',
            'document_date',
            'document_type',
            'receipt_date',
            'delivery_date',
            'created',
            'creator',
            'checked_out',
            'public_trial',
            'file_extension',
            'keywords',
            'filesize',
            'description',
        ]})
        browser.open(view='document_report', data=params)

        workbook = self.load_workbook(browser.contents)
        rows = list(workbook.active.rows)

        expected_titles = [
            u'Title',
            u'Document number',
            u'Last modified',
            u'Author',
            u'Dossier',
            u'Subdossier',
            u'Document-ID',
            u'Document date',
            u'Document type',
            u'Received date',
            u'Sent date',
            u'Created',
            u'Created by',
            u'Checked out by',
            u'Disclosure status',
            u'File extension',
            u'Keywords',
            u'File size',
            u'Description',
        ]
        self.assertEqual(expected_titles, [cell.value for cell in rows[0]])

        expected_values = [
            u'Vertr\xe4gsentwurf',
            u'Client1 1.1 / 1 / 14',
            datetime(2016, 8, 31, 14, 7, 33),
            u'test_user_1_',
            u'Vertr\xe4ge mit der kantonalen Finanzverwaltung',
            None,
            14,
            datetime(2010, 1, 3, 0, 0),
            u'Contract',
            datetime(2010, 1, 3, 0, 0),
            datetime(2010, 1, 3, 0, 0),
            datetime(2016, 8, 31, 14, 7, 33),
            u'robert.ziegler',
            None,
            u'not assessed',
            u'.docx',
            u'Wichtig',
            27413,
            u'Wichtige Vertr\xe4ge',
        ]
        self.assertEqual(expected_values, [cell.value for cell in rows[1]])

    @browsing
    def test_necessary_columns_are_aliased(self, browser):
        self.login(self.regular_user, browser=browser)

        self.checkout_document(self.document)
        self.commit_solr()

        params = self.make_path_param(self.document)

        # The frontend uses these names for the 'checked_out' and
        # 'document_type' columns. They should be aliased accordingly in the
        # DocumentReporter column_settings
        params.update({'columns': [
            'checked_out_fullname',
            'document_type_label',
        ]})
        browser.open(view='document_report', data=params)

        workbook = self.load_workbook(browser.contents)
        rows = list(workbook.active.rows)

        expected_titles = [
            u'Checked out by',
            u'Document type',
        ]
        self.assertEqual(expected_titles, [cell.value for cell in rows[0]])

        expected_values = [
            u'B\xe4rfuss K\xe4thi (kathi.barfuss)',
            u'Contract',
        ]
        self.assertEqual(expected_values, [cell.value for cell in rows[1]])

    @browsing
    def test_supports_custom_fields(self, browser):
        self.login(self.regular_user, browser=browser)

        self.create_propertysheet_for(self.document)
        IDocumentCustomProperties(self.document).custom_properties = {
            "IDocumentMetadata.document_type.offer": {
                "yesorno": True,
                "color": u'Gr\xfcn',
                "colors": [u'Rot', u'Gr\xfcn'],
                "age": 42,
                # Multiline Text fields are not currently indexed in Solr
                "poem": "Lorem\nIpsum",
                "tagline": "Woosh!",
                "birthday": date(2022, 4, 1),
            }
        }
        self.document.reindexObject()
        self.commit_solr()

        params = self.make_path_param(self.document)
        params.update({'columns': [
            'yesorno_custom_field_boolean',
            'color_custom_field_string',
            'colors_custom_field_strings',
            'age_custom_field_int',
            'tagline_custom_field_string',
            'birthday_custom_field_date',
        ]})
        browser.open(view='document_report', data=params)

        workbook = self.load_workbook(browser.contents)
        rows = list(workbook.active.rows)

        expected_titles = [
            u'Yes or no',
            u'Favorite Color',
            u'Colors',
            u'Age',
            u'Tag Line',
            u'Birthday',
        ]
        self.assertEqual(expected_titles, [cell.value for cell in rows[0]])

        expected_values = [
            True,
            u'Gr\xfcn',
            u'Gr\xfcn, Rot',
            42,
            u'Woosh!',
            datetime(2022, 4, 1, 0, 0),
        ]
        self.assertEqual(expected_values, [cell.value for cell in rows[1]])

    @browsing
    def test_supports_multiple_custom_field_with_same_name(self, browser):
        self.login(self.regular_user, browser=browser)

        self.create_propertysheet_for(self.document)
        IDocumentCustomProperties(self.document).custom_properties = {
            "IDocumentMetadata.document_type.offer": {
                "tagline": "Woosh!",
            }
        }

        # add a second additional_title
        create(
            Builder("property_sheet_schema")
            .named("schema2")
            .assigned_to_slots(u"IDocumentMetadata.document_type.report")
            .with_field("textline", u"tagline", u"Tag Line", u"", False))

        self.document.reindexObject()
        self.commit_solr()

        params = self.make_path_param(self.document)
        params.update({'columns': ['title', 'tagline_custom_field_string']})
        browser.open(view='document_report', data=params)

        workbook = self.load_workbook(browser.contents)
        rows = list(workbook.active.rows)

        expected_titles = [u'Title', u'Tag Line']
        self.assertEqual(expected_titles, [cell.value for cell in rows[0]])

        expected_values = [u'Vertr\xe4gsentwurf', u'Woosh!']
        self.assertEqual(expected_values, [cell.value for cell in rows[1]])

    @browsing
    def test_supports_multiple_custom_field_with_different_title(self, browser):
        self.login(self.regular_user, browser=browser)

        create(
            Builder("property_sheet_schema")
            .named("schema1")
            .assigned_to_slots(u"IDocumentMetadata.document_type.offer")
            .with_field("textline", u"tagline", u"Tag Line - 1", u"", False)
        )

        # Another propertysheet for a different slot. We add the same field
        # again but with another title.
        create(
            Builder("property_sheet_schema")
            .named("schema2")
            .assigned_to_slots(u"IDocumentMetadata.document_type.report")
            .with_field("textline", u"tagline", u"Tag Line - 2", u"", False))

        IDocumentMetadata(self.document).document_type = u"offer"
        IDocumentCustomProperties(self.document).custom_properties = {
            "IDocumentMetadata.document_type.offer": {
                "tagline": "Foo",
            }
        }
        self.document.reindexObject()

        IDocumentMetadata(self.subdocument).document_type = u"report"
        IDocumentCustomProperties(self.subdocument).custom_properties = {
            "IDocumentMetadata.document_type.report": {
                "tagline": "Bar",
            }
        }
        self.subdocument.reindexObject()

        self.commit_solr()

        params = self.make_path_param(self.document, self.subdocument)
        params.update({'columns': ['title', 'tagline_custom_field_string']})
        browser.open(view='document_report', data=params)

        workbook = self.load_workbook(browser.contents)
        rows = list(workbook.active.rows)

        expected_titles = [u'Title', u'Tag Line - 2']
        self.assertEqual(expected_titles, [cell.value for cell in rows[0]])

        expected_values = [u'\xdcbersicht der Vertr\xe4ge von 2016', u'Bar']
        self.assertEqual(expected_values, [cell.value for cell in rows[1]])

        expected_values = [u'Vertr\xe4gsentwurf', u'Foo']
        self.assertEqual(expected_values, [cell.value for cell in rows[2]])

    @browsing
    def test_sets_number_format_for_date_fields(self, browser):
        self.login(self.regular_user, browser=browser)

        params = self.make_path_param(self.document)
        params.update({'columns': [
            'changed',
            'document_date',
            'receipt_date',
            'delivery_date',
            'created',
        ]})
        browser.open(view='document_report', data=params)

        workbook = self.load_workbook(browser.contents)
        rows = list(workbook.active.rows)

        expected_values = [
            datetime(2016, 8, 31, 14, 7, 33),
            datetime(2010, 1, 3, 0, 0),
            datetime(2010, 1, 3, 0, 0),
            datetime(2010, 1, 3, 0, 0),
            datetime(2016, 8, 31, 14, 7, 33),
        ]
        self.assertEqual(expected_values, [cell.value for cell in rows[1]])
        self.assertEqual(
            ['DD.MM.YYYY HH:MM', 'DD.MM.YYYY', 'DD.MM.YYYY', 'DD.MM.YYYY', 'DD.MM.YYYY HH:MM'],
            [cell.number_format for cell in rows[1]],
        )

    @browsing
    def test_document_report_with_pseudorelative_path(self, browser):
        self.login(self.regular_user, browser=browser)

        # /ordnungssystem/...
        data = self.make_pseudorelative_path_param(self.document, self.mail_eml)
        data['sort_on'] = 'reference'
        data['sort_order'] = 'desc'
        browser.open(view='document_report', data=data)

        workbook = self.load_workbook(browser.contents)

        # One title row and two data rows
        self.assertEquals(3, len(list(workbook.active.rows)))

        self.assertSequenceEqual(
            [u'Client1 1.1 / 1 / 14',
             14,
             u'Vertr\xe4gsentwurf',
             u'test_user_1_',
             datetime(2010, 1, 3, 0, 0),
             datetime(2010, 1, 3, 0, 0),
             datetime(2010, 1, 3, 0, 0),
             None,
             u'not assessed',
             u'Vertr\xe4ge mit der kantonalen Finanzverwaltung'],
            [cell.value for cell in list(workbook.active.rows)[1]])

        self.assertSequenceEqual(
            [u'Client1 1.1 / 1 / 29',
             29,
             u'Die B\xfcrgschaft',
             u'Freddy H\xf6lderlin <from@example.org>',
             datetime(1999, 1, 1, 0, 0),
             datetime(2016, 8, 31, 0, 0),
             None,
             None,
             u'not assessed',
             u'Vertr\xe4ge mit der kantonalen Finanzverwaltung'],
            [cell.value for cell in list(workbook.active.rows)[2]])

    @browsing
    def test_respects_column_tabbedview_settings_if_exists(self, browser):
        self.login(self.regular_user, browser=browser)

        grid_state = json.dumps(
            {u'sort': {u'field': u'sequence_number', u'direction': u'ASC'},
             u'columns': [
                 {u'width': 30, u'id': u'path_checkbox'},
                 {u'width': 110, u'sortable': True, u'id': u'sequence_number'},
                 {u'width': 200, u'sortable': True, u'id': u'sortable_title'},
                 {u'width': 110, u'sortable': True, u'id': u'document_author'},
                 {u'width': 110, u'sortable': True, u'id': u'document_date'},
                 {u'width': 110, u'hidden': True, u'sortable': True, u'id': u'changed'},
                 {u'width': 110, u'hidden': True, u'sortable': True, u'id': u'created'},
                 {u'width': 110, u'sortable': True, u'id': u'receipt_date'}, {u'width': 110, u'sortable': True, u'id': u'delivery_date'},
                 {u'width': 110, u'sortable': True, u'id': u'checked_out'},
                 {u'width': 110, u'sortable': True, u'hidden': True, u'id': u'public_trial'},
                 {u'width': 110, u'sortable': True, u'id': u'reference'},
                 {u'width': 110, u'sortable': True, u'id': u'file_extension'},
                 {u'width': 110, u'id': u'Subject'},
                 {u'width': 1, u'hidden': True, u'id': u'dummy'}]})

        data = {'view_name': 'mydocuments',
                'gridstate': grid_state}
        browser.open(view='@@tabbed_view/setgridstate', data=data)

        data = self.make_path_param(self.document, self.mail_eml)
        data['view_name'] = 'mydocuments'
        browser.open(view='document_report', data=data)

        workbook = self.load_workbook(browser.contents)

        self.assertEquals(
            [u'Document-ID',
             u'Title',
             u'Author',
             u'Document date',
             u'Received date',
             u'Sent date',
             u'Checked out by',
             u'Document number'],
            [cell.value for cell in list(workbook.active.rows)[0]])

    @browsing
    def test_supports_query_by_listing(self, browser):
        self.login(self.regular_user, browser=browser)

        payload = {'listing_name': 'documents'}

        # Report
        browser.open(view='document_report', data=payload)
        workbook = self.load_workbook(browser.contents)
        rows = list(workbook.active.rows)

        # Listing
        view = '@listing?name=documents&b_size=50'
        browser.open(self.portal, view=view, headers=self.api_headers)

        report_titles = [row[2].value for row in rows[1:]]
        listing_titles = [item.get('title') for item in browser.json.get('items')]

        self.assertEqual(35, len(report_titles))
        self.assertEqual(listing_titles, report_titles)

    @browsing
    def test_supports_query_by_listing_with_filtering(self, browser):
        self.login(self.regular_user, browser=browser)

        payload = {
            'listing_name': 'documents',
            'filters.keywords:record:list': 'Wichtig'
        }

        # Report
        browser.open(view='document_report', data=payload)
        workbook = self.load_workbook(browser.contents)
        rows = list(workbook.active.rows)

        # Listing
        view = '@listing?name=documents&filters.keywords:record:list={}'.format(
            'Wichtig')
        browser.open(self.portal, view=view, headers=self.api_headers)

        report_titles = [row[2].value for row in rows[1:]]
        listing_titles = [item.get('title') for item in browser.json.get('items')]

        self.assertEqual(2, len(report_titles))
        self.assertEqual(listing_titles, report_titles)

    @browsing
    def test_supports_query_by_listing_with_searching(self, browser):
        self.login(self.regular_user, browser=browser)

        payload = {
            'listing_name': 'documents',
            'search': self.document.title
        }

        # Report
        browser.open(view='document_report', data=payload)
        workbook = self.load_workbook(browser.contents)
        rows = list(workbook.active.rows)

        # Listing
        view = '@listing?name=documents&search={}'.format(self.document.title.encode('utf-8'))
        browser.open(self.portal, view=view, headers=self.api_headers)

        report_titles = [row[2].value for row in rows[1:]]
        listing_titles = [item.get('title') for item in browser.json.get('items')]

        self.assertEqual(7, len(report_titles))
        self.assertEqual(listing_titles, report_titles)

    @browsing
    def test_supports_query_by_listing_with_ordering(self, browser):
        self.login(self.regular_user, browser=browser)

        payload = {
            'listing_name': 'documents',
            'sort_order': 'descending',
            'sort_on': 'containing_subdossier'
        }

        # Report
        browser.open(view='document_report', data=payload)
        workbook = self.load_workbook(browser.contents)
        rows = list(workbook.active.rows)

        # Listing sorting descending
        view = '@listing?name=documents&sort_order=descending&sort_on=containing_subdossier&b_size=50'.format(
            self.regular_user.id)
        browser.open(self.portal, view=view, headers=self.api_headers)
        listing_titles_descending = [item.get('title') for item in browser.json.get('items')]

        # Listing sorting ascending
        view = '@listing?name=documents&sort_order=ascending&sort_on=is_subdossier'.format(
            self.regular_user.id)
        browser.open(self.portal, view=view, headers=self.api_headers)
        listing_titles_ascending = [item.get('title') for item in browser.json.get('items')]

        report_titles = [row[2].value for row in rows[1:]]

        self.assertEqual(listing_titles_descending, report_titles)
        self.assertNotEqual(listing_titles_ascending, report_titles)
