from datetime import datetime
from ftw.testbrowser import browsing
from opengever.base.behaviors.lifecycle import ILifeCycle
from opengever.testing import IntegrationTestCase
from openpyxl import load_workbook
from tempfile import NamedTemporaryFile


class TestDispositionExcelExport(IntegrationTestCase):

    @browsing
    def test_label_row(self, browser):
        self.login(self.records_manager, browser)
        browser.open(self.disposition, view='download_excel')

        data = browser.contents
        with NamedTemporaryFile(suffix='.xlsx') as tmpfile:
            tmpfile.write(data)
            tmpfile.flush()
            workbook = load_workbook(tmpfile.name)

            self.assertEquals(
                [u'Reference number', u'Title', u'Start date',
                 u'End date', u'Disclosure status', u'Archival value',
                 u'Comment about archival value assessment',
                 u'Documents', u'Size', u'Appraisal decision'],
                [cell.value for cell in list(workbook.active.rows)[0]])
            self.assertTrue(workbook.active['A1'].font.bold)

    @browsing
    def test_value_rows(self, browser):
        self.login(self.records_manager, browser)

        ILifeCycle(self.offered_dossier_to_destroy).archival_value_annotation = "In Absprache mit ARCH."
        self.offered_dossier_to_archive.public_trial = 'limited-public'

        browser.open(self.disposition, view='download_excel')

        data = browser.contents
        with NamedTemporaryFile(suffix='.xlsx') as tmpfile:
            tmpfile.write(data)
            tmpfile.flush()
            workbook = load_workbook(tmpfile.name)

            rows = list(workbook.active.rows)

            self.assertEquals(
                [u'Client1 1.1 / 12', u'Hannah Baufrau', datetime(2000, 1, 1, 0, 0),
                 datetime(2000, 1, 31, 0, 0), u'public (with restrictions)',
                 u'archival worthy', None, 9, u'1 KB', u'archival worthy'],
                [cell.value for cell in rows[1]])

            self.assertEquals(
                [u'Client1 1.1 / 14', u'Hans Baumann', datetime(2000, 1, 1, 0, 0),
                 datetime(2000, 1, 15, 0, 0), u'not assessed',
                 u'not archival worthy', u'In Absprache mit ARCH.',
                 9, u'1 KB', u'not archival worthy'],
                [cell.value for cell in rows[2]])

    @browsing
    def test_file_name(self, browser):
        self.enable_languages()

        self.login(self.records_manager, browser)

        browser.open()
        browser.click_on("Deutsch")
        browser.open(self.disposition, view='download_excel')
        fname = eval(browser.headers.get(
            "content-disposition").split(";")[1].split("=")[1])
        expected = "bewertungsliste_angebot-31-8-2016.xlsx"
        self.assertEquals(expected, fname)

        browser.open()
        browser.click_on(u'Fran\xe7ais')
        browser.open(self.disposition, view='download_excel')
        fname = eval(browser.headers.get(
            "content-disposition").split(";")[1].split("=")[1])
        expected = "liste_evaluations_angebot-31-8-2016.xlsx"
        self.assertEquals(expected, fname)

    @browsing
    def test_escape_invalid_chars_in_sheet_title(self, browser):
        self.login(self.records_manager, browser)
        self.disposition.title = u'Angebot 06.04.2023: Rechtsdienst'
        browser.open(self.disposition, view='download_excel')

        data = browser.contents
        with NamedTemporaryFile(suffix='.xlsx') as tmpfile:
            tmpfile.write(data)
            tmpfile.flush()
            workbook = load_workbook(tmpfile.name)

            self.assertEquals([u'Angebot 06.04.2023 Rechtsdien'],
                              [sheet.title for sheet in workbook.worksheets])
