from DateTime import DateTime
from math import floor
from Missing import Value as MissingValue
from opengever.base.solr import OGSolrContentListingObject
from opengever.base.solr.fields import DateListingField
from opengever.ogds.base.actor import Actor
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from plone.api.portal import get_localized_time
from StringIO import StringIO
from zope.globalrequest import getRequest
from zope.i18n import translate
import dateutil
import six


DATE_NUMBER_FORMAT = 'DD.MM.YYYY'
DATETIME_NUMBER_FORMAT = 'DD.MM.YYYY HH:MM'


def value(input_string):
    """Return unicode"""

    if not isinstance(input_string, unicode):
        input_string = input_string.decode('utf-8')
    return input_string


def readable_actor(actorid):
    """Helper method which returns the actor label
    instead of the actorid"""
    if actorid is None:
        return ''

    actor = Actor.lookup(six.ensure_text(actorid))
    if actor.actor_type == 'null':
        return actorid or ''
    return actor.get_label()


def readable_date(date):
    """Helper method to return a localized date"""
    return get_localized_time(date) if date else None


def heuristic_for_11pt_calibri(charcount):
    # Adapted from:
    # https://msdn.microsoft.com/en-us/library/documentformat.openxml.spreadsheet.column.aspx
    max_digit_width = 7.0
    cell_padding = 5.0
    return floor((charcount * max_digit_width + cell_padding) / max_digit_width * 256) / 256


def approximate_column_width(cells):
    return heuristic_for_11pt_calibri(max(
        len(unicode(cell.value)) if cell.value else 0
        for cell in cells
    ))


class StringTranslater(object):
    """provide the translate method as helper method
    for the given domain and request"""

    def __init__(self, request, domain):
        self.request = request
        self.domain = domain

    def translate(self, value):
        """ return the translated string"""

        if self.request is None:
            self.request = getRequest()

        if value:
            return translate(value, domain=self.domain, context=self.request)
        return None


class XLSReporter(object):
    """XLS Reporter View generates a xls-report for the given results set.
    """

    def __init__(self, request, attributes, results,
                 sheet_title=u' ', footer=u'', portrait_format=False,
                 blank_header_rows=0, field_mapper=None):
        """Initalize the XLS reporter
        Arguments:
        attributes -- a list of mappings (with 'id', 'title', 'transform')
        results -- a list of objects, brains or sqlalchemy objects
        """

        self.attributes = attributes
        self.results = results
        self.request = request
        self.sheet_title = sheet_title
        self.footer = footer
        self.portrait_format = portrait_format
        self.blank_header_rows = blank_header_rows
        self.field_mapper = field_mapper

    def __call__(self):
        workbook = self.prepare_workbook()
        # save the Workbook-data in to a StringIO
        data = StringIO()
        workbook.save(data)
        data.seek(0)
        return data.read()

    def prepare_workbook(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = self.sheet_title
        sheet.oddFooter.center.text = self.footer

        self.insert_label_row(sheet)
        self.insert_value_rows(sheet)

        column_widths = {
            get_column_letter(i + 1): approximate_column_width(col)
            for i, col in enumerate(sheet.columns)
        }

        for index, width in column_widths.iteritems():
            sheet.column_dimensions[index].width = width

        return workbook

    def insert_label_row(self, sheet):
        title_font = Font(bold=True)
        for i, attr in enumerate(self.attributes, 1):
            cell = sheet.cell(row=self.blank_header_rows + 1, column=i)
            cell.value = translate(attr.get('title', ''), context=self.request)
            cell.font = title_font

    def insert_value_rows(self, sheet):
        for row, obj in enumerate(self.results, 2):
            for column, attr in enumerate(self.attributes, 1):
                cell = sheet.cell(row=self.blank_header_rows + row, column=column)

                value = self.get_value(obj, attr)
                cell.value = value

                # Enable text wrapping for cells that may contain multiline text
                cell.alignment = Alignment(wrap_text=True)

                if 'hyperlink' in attr:
                    cell.hyperlink = attr.get('hyperlink')(value, obj)
                    cell.style = 'Hyperlink'

                if 'number_format' in attr:
                    cell.number_format = attr.get('number_format')

                if 'fold_by_method' in attr:
                    sheet.row_dimensions[cell.row].outlineLevel = attr.get('fold_by_method')(value)

    def get_value(self, obj, attr):
        if isinstance(obj, dict):
            value = obj.get(attr.get('id'), attr.get('default'))

        elif isinstance(obj, OGSolrContentListingObject):
            field_name = attr.get('id')
            field = self.field_mapper.get(field_name)

            value = field.get_value(obj)

            # Date handling will need to be revisited once this is rebased
            # onto the optimize-reindex-of-contained-objects branch
            if isinstance(value, DateTime):
                value = value.asdatetime()
            elif isinstance(field, DateListingField) and value:
                value = dateutil.parser.parse(value)

        else:
            if 'default' in attr:
                value = getattr(obj, attr.get('id'), attr.get('default'))
            else:
                value = getattr(obj, attr.get('id'))

        if attr.get('callable'):
            value = value()
        if attr.get('transform'):
            value = attr.get('transform')(value)
        if value == MissingValue:
            value = ''
        if isinstance(value, (list, set, tuple)):
            value = ', '.join(value)

        return value
