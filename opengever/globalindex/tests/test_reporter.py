from datetime import datetime
from ftw.testbrowser import browsing
from opengever.testing import IntegrationTestCase
from openpyxl import load_workbook
from tempfile import NamedTemporaryFile
import json


class TestTaskReporter(IntegrationTestCase):

    @browsing
    def test_task_report(self, browser):
        self.login(self.regular_user, browser=browser)
        browser.open(self.task, method='PATCH', headers=self.api_headers,
                     data=json.dumps({'text': 'A description'}))

        browser.open(view='task_report',
                     data={'view_name': 'mytasks',
                           'task_ids': [self.task.get_sql_object().task_id,
                                        self.meeting_task.get_sql_object().task_id,
                                        self.subtask.get_sql_object().task_id]})

        with NamedTemporaryFile(delete=False, suffix='.xlsx') as tmpfile:
            tmpfile.write(browser.contents)
            tmpfile.flush()
            workbook = load_workbook(tmpfile.name)

        # labels
        self.assertSequenceEqual(
            [u'Title', u'State',
             u'Deadline', u'Completed on',
             u'Dossier', u'Main task', u'Issuer',
             u'Issuing organization', u'Responsible',
             u'Task type', 'Tenant', u'Sequence number',
             u'Description'],
            [cell.value for cell in list(workbook.active.rows)[0]])

        # self.task
        self.assertSequenceEqual(
            [self.task.title,  u'In progress',
             datetime(2016, 11, 1, 0, 0), None,
             self.dossier.title, None, u'Ziegler Robert (robert.ziegler)',
             u'Finanz\xe4mt', u'B\xe4rfuss K\xe4thi (kathi.barfuss)',
             u'For your review', u'plone', 1, u'A description'],
            [cell.value for cell in list(workbook.active.rows)[1]])

        # self.meeting_task
        self.assertSequenceEqual(
            [self.meeting_task.title, u'In progress',
             datetime(2016, 11, 1, 0, 0), None,
             self.meeting_dossier.title, None, u'Ziegler Robert (robert.ziegler)',
             u'Finanz\xe4mt', u'Ziegler Robert (robert.ziegler)',
             u'For your review', u'plone', 9, None],
            [cell.value for cell in list(workbook.active.rows)[2]])

        # self.subtask
        self.assertSequenceEqual(
            [self.subtask.title,  u'Resolved',
             datetime(2016, 11, 1, 0, 0), None,
             self.dossier.title, self.task.title, u'Ziegler Robert (robert.ziegler)',
             u'Finanz\xe4mt', u'B\xe4rfuss K\xe4thi (kathi.barfuss)',
             u'For your review', u'plone', 2, None],
            [cell.value for cell in list(workbook.active.rows)[3]])

    @browsing
    def test_respects_column_tabbedview_settings_if_exists(self, browser):
        self.login(self.regular_user, browser=browser)

        # Changing tabbedviews grid state (order and visible columns)
        data = {'view_name': 'mytasks',
                'gridstate': json.dumps({
                    u'sort': {u'field': u'modified', u'direction': u'ASC'},
                    u'columns': [
                        {u'width': 30, u'id': u'task_id_checkbox_helper'},
                        {u'width': 110, u'sortable': True, u'id': u'sequence_number'},
                        {u'width': 110, u'sortable': True, u'id': u'review_state'},
                        {u'width': 110, u'sortable': True, u'id': u'responsible'},
                        {u'width': 110, u'sortable': True, u'id': u'issuer'},
                        {u'width': 110, u'sortable': True, u'id': u'task_type'},
                        {u'width': 110, u'sortable': True, u'id': u'title'},
                        {u'width': 110, u'hidden': True, u'sortable': True, u'id': u'deadline'},
                        {u'width': 110, u'hidden': True, u'sortable': True, u'id': u'completed'},
                        {u'width': 110, u'hidden': True, u'sortable': True, u'id': u'created'},
                        {u'width': 110, u'sortable': True, u'id': u'containing_dossier'},
                        {u'width': 110, u'sortable': True, u'id': u'issuing_org_unit'},
                        {u'width': 1, u'hidden': True, u'id': u'dummy'}]})}

        browser.open(view='@@tabbed_view/setgridstate', data=data)

        browser.open(view='task_report',
                     data={'view_name': 'mytasks',
                           'task_ids': [self.task.get_sql_object().task_id,
                                        self.meeting_task.get_sql_object().task_id]})

        with NamedTemporaryFile(delete=False, suffix='.xlsx') as tmpfile:
            tmpfile.write(browser.contents)
            tmpfile.flush()
            workbook = load_workbook(tmpfile.name)

        # labels
        self.assertSequenceEqual(
            [u'Sequence number', u'State',
             u'Responsible', u'Issuer',
             u'Task type', u'Title',
             u'Dossier'],
            [cell.value for cell in list(workbook.active.rows)[0]])

    @browsing
    def test_task_report_by_task_ressource_id(self, browser):
        self.login(self.regular_user, browser=browser)

        browser.open(view='task_report',
                     data={'view_name': 'mytasks',
                           'tasks': [self.task.absolute_url(),
                                     self.meeting_task.absolute_url()]})

        with NamedTemporaryFile(delete=False, suffix='.xlsx') as tmpfile:
            tmpfile.write(browser.contents)
            tmpfile.flush()
            workbook = load_workbook(tmpfile.name)

        # self.task
        task_cells = list(workbook.active.rows)[1]
        self.assertSequenceEqual(
            [self.task.title,  u'In progress',
             datetime(2016, 11, 1, 0, 0), None,
             self.dossier.title, None, u'Ziegler Robert (robert.ziegler)',
             u'Finanz\xe4mt', u'B\xe4rfuss K\xe4thi (kathi.barfuss)',
             u'For your review', u'plone', 1, None],
            [cell.value for cell in task_cells])
        self.assertEqual(self.task.absolute_url(), task_cells[0].hyperlink.target)

        # self.meeting_task
        self.assertSequenceEqual(
            [self.meeting_task.title, u'In progress',
             datetime(2016, 11, 1, 0, 0), None,
             self.meeting_dossier.title, None, u'Ziegler Robert (robert.ziegler)',
             u'Finanz\xe4mt', u'Ziegler Robert (robert.ziegler)',
             u'For your review', u'plone', 9, None],
            [cell.value for cell in list(workbook.active.rows)[2]])

    @browsing
    def test_include_subtasks(self, browser):
        self.login(self.regular_user, browser=browser)

        browser.open(view='task_report',
                     data={'view_name': 'mytasks',
                           'include_subtasks': True,
                           'tasks': [self.task.absolute_url(),
                                     self.meeting_task.absolute_url()]})

        with NamedTemporaryFile(delete=False, suffix='.xlsx') as tmpfile:
            tmpfile.write(browser.contents)
            tmpfile.flush()
            workbook = load_workbook(tmpfile.name)

        # self.task
        self.assertSequenceEqual(
            [self.task.title, u'In progress',
             datetime(2016, 11, 1, 0, 0), None,
             self.dossier.title, None, u'Ziegler Robert (robert.ziegler)',
             u'Finanz\xe4mt', u'B\xe4rfuss K\xe4thi (kathi.barfuss)',
             u'For your review', u'plone', 1, None],
            [cell.value for cell in list(workbook.active.rows)[1]])

        # self.subtask
        self.assertSequenceEqual(
            [self.subtask.title,  u'Resolved',
             datetime(2016, 11, 1, 0, 0), None,
             self.dossier.title, self.task.title, u'Ziegler Robert (robert.ziegler)',
             u'Finanz\xe4mt', u'B\xe4rfuss K\xe4thi (kathi.barfuss)',
             u'For your review', u'plone', 2, None],
            [cell.value for cell in list(workbook.active.rows)[2]])

        # self.meeting_task
        self.assertSequenceEqual(
            [self.meeting_task.title, u'In progress',
             datetime(2016, 11, 1, 0, 0), None,
             self.meeting_dossier.title, None, u'Ziegler Robert (robert.ziegler)',
             u'Finanz\xe4mt', u'Ziegler Robert (robert.ziegler)',
             u'For your review', u'plone', 9, None],
            [cell.value for cell in list(workbook.active.rows)[3]])

        # self.meeting_subtask
        self.assertSequenceEqual(
            [self.meeting_subtask.title, u'Resolved',
             datetime(2016, 11, 1, 0, 0), None,
             self.meeting_dossier.title, self.meeting_task.title,
             u'Ziegler Robert (robert.ziegler)',
             u'Finanz\xe4mt', u'Ziegler Robert (robert.ziegler)',
             None, u'plone', 10, None],
            [cell.value for cell in list(workbook.active.rows)[4]])

    @browsing
    def test_does_not_include_subtasks(self, browser):
        self.login(self.regular_user, browser=browser)

        browser.open(view='task_report',
                     data={'view_name': 'mytasks',
                           'include_subtasks': False,
                           'tasks': [self.task.absolute_url(),
                                     self.meeting_task.absolute_url()]})

        with NamedTemporaryFile(delete=False, suffix='.xlsx') as tmpfile:
            tmpfile.write(browser.contents)
            tmpfile.flush()
            workbook = load_workbook(tmpfile.name)

        # self.task
        task_cells = list(workbook.active.rows)[1]
        self.assertSequenceEqual(
            [self.task.title,  u'In progress',
             datetime(2016, 11, 1, 0, 0), None,
             self.dossier.title, None, u'Ziegler Robert (robert.ziegler)',
             u'Finanz\xe4mt', u'B\xe4rfuss K\xe4thi (kathi.barfuss)',
             u'For your review', u'plone', 1, None],
            [cell.value for cell in task_cells])
        self.assertEqual(self.task.absolute_url(), task_cells[0].hyperlink.target)

        # self.meeting_task
        self.assertSequenceEqual(
            [self.meeting_task.title, u'In progress',
             datetime(2016, 11, 1, 0, 0), None,
             self.meeting_dossier.title, None, u'Ziegler Robert (robert.ziegler)',
             u'Finanz\xe4mt', u'Ziegler Robert (robert.ziegler)',
             u'For your review', u'plone', 9, None],
            [cell.value for cell in list(workbook.active.rows)[2]])


class TestUserReporter(IntegrationTestCase):

    @browsing
    def test_empty_users_report(self, browser):
        self.login(self.administrator, browser=browser)
        browser.open(view='user_report', data={'user_ids': []})

        self.assertEquals('Error You have not selected any items.',
                          browser.css('.portalMessage.error').text[0])

    @browsing
    def test_download_users_xlsx(self, browser):
        self.login(self.administrator, browser=browser)
        user_ids = {'user_ids': [self.regular_user.id, self.administrator.id]}

        browser.open(view='user_report', data=user_ids)
        self.assertEqual(browser.status_code, 200)
        with NamedTemporaryFile(delete=False, suffix='.xlsx') as tmpfile:
            tmpfile.write(browser.contents)
            tmpfile.flush()
            workbook = load_workbook(tmpfile.name)

        task_cells = list(workbook.active.rows)

        # if ror_num != 0 will remove the table header
        cell_values = [[cell.value for cell in row] for row_num, row in enumerate(task_cells) if row_num != 0]
        expected_values = [
            [
                u'kathi.barfuss',
                True,
                u'K\xe4thi',
                u'B\xe4rfuss',
                u'B\xe4rfuss K\xe4thi',
                u'Staatsarchiv',
                u'Arch',
                u'Staatskanzlei',
                u'SK',
                None,
                u'foo@example.com',
                u'bar@example.com',
                u'http://www.example.com',
                u'012 34 56 78',
                u'012 34 56 77',
                u'012 34 56 76',
                u'Frau',
                u'Gesch\xe4ftsf\xfchrerin',
                u'nix',
                u'Kappelenweg 13',
                u'Postfach 1234',
                u'1234',
                u'Vorkappelen',
                u'Schweiz',
                None
            ],

            [
                u'nicole.kohler',
                True, u'Nicole',
                u'Kohler',
                u'Kohler Nicole',
                None,
                None,
                None,
                None,
                None,
                u'nicole.kohler@gever.local',
                None, None, None, None, None,
                None, None, None, None, None,
                None, None, None, None,
            ]
        ]
        self.assertSequenceEqual(expected_values, cell_values)


class TestOGDSGroupsMembershipReporter(IntegrationTestCase):

    @browsing
    def test_groups_membership_report(self, browser):
        # Login as an administrator to access the report
        self.login(self.administrator, browser=browser)

        browser.open(view='group-memberships-report')

        # Ensure the response status is 200 (OK)
        self.assertEqual(browser.status_code, 200)

        # Save the returned Excel content to a temporary file
        with NamedTemporaryFile(delete=False, suffix='.xlsx') as tmpfile:
            tmpfile.write(browser.contents)
            tmpfile.flush()
            workbook = load_workbook(tmpfile.name)

        # Check the headers/columns in the Excel sheet
        self.assertSequenceEqual(
            ['Username', 'User-ID', 'Group name', 'Group-ID'],
            [cell.value for cell in list(workbook.active.rows)[0]]
        )

        # Get the rows (skip the first row which contains headers)
        rows = list(workbook.active.rows)[1:]
        expected_values = [
            [u'kathi.barfuss', u'regular_user', u'fa_users', u'fa_users'],
            [u'jurgen.fischer', u'archivist', u'fa_users', u'fa_users'],
            [u'ramon.flucht', u'records_manager', u'fa_users', u'fa_users'],
            [u'gunther.frohlich', u'gunther.frohlich', u'fa_users', u'fa_users'],
            [u'faivel.fruhling', u'dossier_manager', u'fa_users', u'fa_users'],
            [u'fridolin.hugentobler', u'fridolin.hugentobler', u'fa_users', u'fa_users'],
            [u'maja.harzig', u'limited_admin', u'fa_users', u'fa_users'],
            [u'herbert.jager', u'meeting_user', u'fa_users', u'fa_users'],
            [u'nicole.kohler', u'nicole.kohler', u'fa_users', u'fa_users'],
            [u'jurgen.konig', u'jurgen.konig', u'fa_users', u'fa_users'],
            [u'propertysheets.manager', u'propertysheets_manager', u'fa_users', u'fa_users'],
            [u'webaction.manager', u'webaction_manager', u'fa_users', u'fa_users'],
            [u'david.meier', u'member_admin', u'fa_users', u'fa_users'],
            [u'franzi.muller', u'committee_responsible', u'fa_users', u'fa_users'],
            [u'hans.peter', u'hans.peter', u'fa_users', u'fa_users'],
            [u'beatrice.schrodinger', u'beatrice.schrodinger', u'fa_users', u'fa_users'],
            [u'committee.secretary', u'committee.secretary', u'fa_users', u'fa_users'],
            [u'service.user', u'service_user', u'fa_users', u'fa_users'],
            [u'robert.ziegler', u'robert.ziegler', u'fa_users', u'fa_users'],
            [u'kathi.barfuss', u'regular_user', u'projekt_a', u'projekt_a'],
            [u'robert.ziegler', u'robert.ziegler', u'projekt_a', u'projekt_a'],
            [u'james.bond', u'james.bond', u'rk_users', u'rk_users'],
            [u'herbert.jager', u'meeting_user', u'projekt_b', u'projekt_b'],
            [u'franzi.muller', u'committee_responsible', u'projekt_b', u'projekt_b'],
            [u'nicole.kohler', u'nicole.kohler', u'committee_rpk_group', u'committee_rpk_group'],
            [u'franzi.muller', u'committee_responsible', u'committee_rpk_group', u'committee_rpk_group'],
            [u'nicole.kohler', u'nicole.kohler', u'committee_ver_group', u'committee_ver_group'],
            [u'franzi.muller', u'committee_responsible', u'committee_ver_group', u'committee_ver_group'],
            [u'jurgen.konig', u'jurgen.konig', u'fa_inbox_users', u'fa_inbox_users']
        ]

        for row, expected_row in zip(rows, expected_values):
            self.assertSequenceEqual(
                expected_row,
                [cell.value for cell in row]
            )
