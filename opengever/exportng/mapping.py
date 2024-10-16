# -*- coding: utf-8 -*-
from openpyxl import load_workbook
import itertools
import os.path


def read_mapping(mapping_file):
    filename = os.path.join(
        os.path.dirname(__file__), mapping_file)
    wb = load_workbook(filename=filename)
    sheet = wb.active
    rows = []
    colnames = [cell.value for cell in sheet[1]]
    for row in sheet.iter_rows(min_row=2):
        data = {k: row[idx].value for idx, k in enumerate(colnames)}
        rows.append(data)

    mappings = {}
    for k, v in itertools.groupby(rows, lambda x: x['table_name']):
        mappings[k] = list(v)
    return mappings
