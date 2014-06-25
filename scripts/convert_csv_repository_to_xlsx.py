""" This script converts a repository.csv file back to a xlsx file.

!! THIS SCRIPT NEEDS OPENPYXL WHICH IS NOT INCLUDED BY DEFAULT !!

Usage: bin/zopepy scripts/convert_csv_repository_to_xlsx.py <path to repository csv file> <path for new xlsx file>

The examplecontent.xlsx serves as template for header and formatting.
The header is hardcoaded to be 5 rows high and the 5th row contains
the column headers.
"""
from openpyxl import load_workbook
import csv
import os
import sys


def main():
    if len(sys.argv) != 3:
        raise Exception('Please use with the following parameter: '
                        '<path to repository csv file> '
                        '<path for new xlsx file>')

    csv_path = sys.argv[1]
    xlsx_path = sys.argv[2]
    xlsx_template = 'opengever/examplecontent/data/examplecontent.xlsx'

    if not os.path.exists(csv_path):
        raise Exception('Could not find csv file at %s.' % csv_path)
    if os.path.exists(xlsx_path):
        raise Exception('File already exists at %s.' % xlsx_path)

    workbook = load_workbook(xlsx_template)
    worksheet = workbook.worksheets[0]

    # clear example content
    for row in worksheet.rows[5:]:
        for cell in row:
            cell.value = None
    worksheet.garbage_collect()

    # mapping in which column which data is (title, reference_number, etc.)
    column_mapping = {}
    for col_nr, key_cell in enumerate(worksheet.rows[4]):
        if key_cell.value is None:
            continue
        column_mapping[key_cell.value] = col_nr

    # write data from csv
    with open(csv_path, 'rb') as csvfile:
        csvreader = csv.reader(csvfile)
        csvkeys = next(csvreader)  # skip header

        for row_nr, csvline in enumerate(csvreader, start=5):
            for col_nr, csvcell in enumerate(csvline):

                # resolve column number by header
                if not csvkeys[col_nr] in column_mapping:
                    continue
                final_col_nr = column_mapping[csvkeys[col_nr]]

                # write data
                wscell = worksheet.cell(row=row_nr+1,
                                        column=final_col_nr+1)
                wscell.value = csvcell

    workbook.save(xlsx_path)


if __name__ == '__main__':
    main()
